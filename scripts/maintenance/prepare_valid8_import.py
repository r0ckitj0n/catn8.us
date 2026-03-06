#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
import tempfile
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

INACTIVE_KEYWORDS = (
    "inactive",
    "old",
    "archive",
    "archived",
    "expired",
    "disabled",
    "previous",
    "legacy",
)

HEADER_ALIASES = {
    "title": ("title", "service", "site", "account", "name", "app"),
    "url": ("url", "website", "web", "login url", "site url"),
    "username": ("username", "user", "login", "email", "userid", "user name"),
    "password": ("password", "pass", "passwd", "pw", "pwd"),
    "notes": ("notes", "note", "recovery", "recovery code", "codes", "comments", "memo"),
    "category": ("category", "type", "group", "folder", "section"),
    "is_active": ("active", "is_active", "status", "enabled"),
    "is_favorite": ("favorite", "favourite", "star", "is_favorite"),
    "password_strength": ("password strength", "strength", "security"),
    "last_changed_at": ("last changed", "changed", "updated", "last updated", "modified", "date changed"),
}


def normalize_header(value: str) -> str:
    value = (value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return value.strip()


def detect_columns(headers: List[str]) -> Dict[str, int]:
    normalized = [normalize_header(h) for h in headers]
    out: Dict[str, int] = {}
    for target, aliases in HEADER_ALIASES.items():
        for idx, h in enumerate(normalized):
            if h in aliases:
                out[target] = idx
                break
            for alias in aliases:
                if alias in h and target not in out:
                    out[target] = idx
    return out


def parse_bool(value: str, default: bool) -> bool:
    v = (value or "").strip().lower()
    if not v:
        return default
    if v in ("1", "true", "yes", "y", "active", "enabled", "current"):
        return True
    if v in ("0", "false", "no", "n", "inactive", "disabled", "old", "archived"):
        return False
    return default


def sheet_implies_inactive(sheet_name: str) -> bool:
    s = (sheet_name or "").strip().lower()
    if any(k in s for k in INACTIVE_KEYWORDS):
        return True

    # Tabs named with historical snapshots (e.g. "Jon 2017-01-19") are treated as inactive archives.
    if re.search(r"(19|20)\d{2}[-_/\.](0[1-9]|1[0-2])[-_/\.](0[1-9]|[12]\d|3[01])", s):
        return True

    return False


def row_to_record(row: List[str], cols: Dict[str, int], default_category: str, default_active: bool, source_doc: str, source_tab: str) -> Optional[Dict[str, object]]:
    def get(name: str) -> str:
        i = cols.get(name)
        if i is None or i >= len(row):
            return ""
        return (row[i] or "").strip()

    title = get("title")
    url = get("url")
    username = get("username")
    password = get("password")
    if not title and url:
        title = url

    if not title or not username or not password:
        return None

    category = get("category") or default_category or "Imported"
    notes = get("notes")
    active = parse_bool(get("is_active"), default_active)

    rec: Dict[str, object] = {
        "title": title,
        "url": url or None,
        "username": username,
        "password": password,
        "notes": notes or None,
        "category": category,
        "is_active": 1 if active else 0,
        "is_favorite": 1 if parse_bool(get("is_favorite"), False) else 0,
        "password_strength": get("password_strength") or "",
        "last_changed_at": get("last_changed_at") or "",
        "source_document": source_doc,
        "source_tab": source_tab,
    }
    return rec


def parse_csv_bytes(data: bytes, source_doc: str, tab_name: str) -> List[Dict[str, object]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = rows[0]
    cols = detect_columns(headers)
    inactive = sheet_implies_inactive(tab_name)
    records: List[Dict[str, object]] = []
    for raw in rows[1:]:
        rec = row_to_record(raw, cols, tab_name, not inactive, source_doc, tab_name)
        if rec:
            records.append(rec)
    return records


def parse_xlsx(path: Path, source_doc: str) -> List[Dict[str, object]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    all_records: List[Dict[str, object]] = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = ["" if v is None else str(v) for v in rows[0]]
        cols = detect_columns(headers)
        inactive = sheet_implies_inactive(ws.title)
        for values in rows[1:]:
            raw = ["" if v is None else str(v) for v in values]
            rec = row_to_record(raw, cols, ws.title, not inactive, source_doc, ws.title)
            if rec:
                all_records.append(rec)
    return all_records


def is_google_sheet_url(url: str) -> Optional[str]:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None


def is_google_drive_file_url(url: str) -> Optional[str]:
    m = re.search(r"/file/d/([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None


def download_url(url: str) -> Tuple[bytes, str]:
    req = urllib.request.Request(url, headers={"User-Agent": "catn8-valid8-import/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
        ctype = (resp.headers.get("Content-Type") or "").lower()
    return data, ctype


def fetch_source(source: str) -> Tuple[bytes, str, str]:
    if re.match(r"^https?://", source, flags=re.I):
        sheet_id = is_google_sheet_url(source)
        if sheet_id:
            export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
            data, ctype = download_url(export_url)
            return data, ctype, f"google-sheet:{sheet_id}"

        file_id = is_google_drive_file_url(source)
        if file_id:
            export_url = f"https://drive.google.com/uc?export=download&id={file_id}"
            data, ctype = download_url(export_url)
            return data, ctype, f"google-drive:{file_id}"

        data, ctype = download_url(source)
        return data, ctype, source

    path = Path(source)
    data = path.read_bytes()
    return data, "", str(path)


def parse_source(source: str) -> List[Dict[str, object]]:
    data, ctype, source_name = fetch_source(source)

    lower_source = source.lower()
    is_xlsx = (
        lower_source.endswith(".xlsx")
        or "spreadsheetml" in ctype
        or data[:2] == b"PK"
    )

    if is_xlsx:
        with tempfile.NamedTemporaryFile(prefix="valid8-", suffix=".xlsx", delete=False) as tmp:
            tmp.write(data)
            tmp_path = Path(tmp.name)
        try:
            return parse_xlsx(tmp_path, source_name)
        finally:
            try:
                tmp_path.unlink(missing_ok=True)
            except OSError:
                pass

    head = data[:40000].lower()
    if ("html" in ctype and (b"servicelogin" in head or b"accounts.google.com" in head)) or (
        head.startswith(b"<!doctype html") and (b"servicelogin" in head or b"accounts.google.com" in head)
    ):
        raise RuntimeError(f"Source requires login and is not directly readable: {source}")

    return parse_csv_bytes(data, source_name, "Imported")


def normalize_strength(value: object) -> int:
    s = str(value or "").strip()
    if not s:
        return 1
    try:
        n = int(float(s))
    except Exception:
        n = 1
    return max(1, min(5, n))


def normalize_date(value: object) -> Optional[str]:
    s = str(value or "").strip()
    return s or None


def dedupe_records(records: List[Dict[str, object]]) -> List[Dict[str, object]]:
    seen = set()
    out: List[Dict[str, object]] = []
    for r in records:
        key = (
            str(r.get("title") or "").strip().lower(),
            str(r.get("url") or "").strip().lower(),
            str(r.get("username") or "").strip().lower(),
            str(r.get("password") or ""),
            str(r.get("notes") or ""),
        )
        if key in seen:
            continue
        seen.add(key)
        r["password_strength"] = normalize_strength(r.get("password_strength"))
        r["last_changed_at"] = normalize_date(r.get("last_changed_at"))
        out.append(r)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Prepare VALID8 import rows from spreadsheet sources")
    ap.add_argument("--source", action="append", required=True, help="Source file path or URL (repeatable)")
    ap.add_argument("--output", default=".local/state/valid8/import_rows.json", help="Output JSON file")
    args = ap.parse_args()

    records: List[Dict[str, object]] = []
    for src in args.source:
        parsed = parse_source(src)
        records.extend(parsed)

    records = dedupe_records(records)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

    active = sum(1 for r in records if int(r.get("is_active") or 0) == 1)
    inactive = len(records) - active
    print(f"Prepared {len(records)} rows (active={active}, inactive={inactive}) -> {out_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
